# sübhan akbenli
# eğer adam harcamayı listelemeden önce o harcama içeriğini veritabanından silerse o zaman uyarı verip silmesin ya da silmek istediğini sorsun
import sqlite3
from PyQt5.QtWidgets import *
import sys
from datetime import datetime,timedelta
from ui_designs.rezTakip_ui import *
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QCursor
from PyQt5.QtGui import QColor

from ui_designs.musteriHesap_ui import *
TABLO_AKTIVITELER="aktiviteler"
INSERT_AKTIVITELER="(aktivite,TL,DOLAR,EURO,KART)"

TABLO_REZERVASYONLAR="rezervasyonlar"
INSERT_REZERVASYONLAR="(aktivite,otelAdi,adSoyad,rezDate,telefon,fiyat,paraBirimi)"

TABLO_MUSTERILER="musteriler"
INSERT_MUSTERILER="(otelAdi,telefon)"

TABLO_HARCAMA="harcama"
INSERT_HARCAMA="(rezId,otelAdi,adSoyad,harcamatarihi,paketAdi,tutar,odendi)"

TABLO_PAKETLER="paketler"
INSERT_PAKETLER="(paketAdi,tutar)"
Uygulama = QApplication(sys.argv)
rezTakip_main_window = QMainWindow()
rezTakip_ui = Ui_rezTakip_MainWindow()
rezTakip_ui.setupUi(rezTakip_main_window)
rezTakip_ui.statusbar.setStyleSheet("font: 75 14pt 'Times New Roman'; background-color:rgb(200,200,200);color:rgb(0,0,0)")



from PyQt5.QtWidgets import QMenu, QTableWidgetItem

# Bu kısmı ilgili class veya fonksiyon içinde kullanabilirsiniz.

# TableWidget'e sağ tık menüsü ekleme

# Sağ tık olayını bağlamak için

class CustomSpinBox(QDoubleSpinBox):
    def __init__(self):
        super().__init__()
        self.setRange(0, 9999999)

    def wheelEvent(self, event):
        # Fare tekerleği olaylarını engelle
        event.ignore()

class CustomComboBox(QComboBox):
    def wheelEvent(self, event):
        event.ignore()

class CustomQDateEdit(QDateEdit):
    def wheelEvent(self, event):
        event.ignore()

class rezTakip():

    def __init__(self):
        self.conn=sqlite3.connect("database.db")
        self.curs=self.conn.cursor()
        self.curs.execute("CREATE TABLE IF NOT EXISTS aktiviteler (aktivite Text PRIMARY KEY,TL FLOAT,DOLAR FLOAT,EURO FLOAT,KART FLOAT)")

        self.curs.execute("CREATE TABLE IF NOT EXISTS rezervasyonlar \
                        (rezID INTEGER  PRIMARY KEY AUTOINCREMENT,aktivite Text ,otelAdi Text,\
                            adSoyad TEXT,rezDate DATE,telefon TEXT,fiyat FLOAT,paraBirimi text,ekstralar FLOAT)")

        self.curs.execute("CREATE TABLE IF NOT EXISTS musteriler \
                        (otelAdi Text PRIMARY KEY,telefon Text)")

        self.curs.execute("CREATE TABLE IF NOT EXISTS harcama \
            (rezId INT,HarcamaId INTEGER PRIMARY KEY AUTOINCREMENT,otelAdi Text,\
                            adSoyad TEXT,harcamatarihi Date,paketAdi Text,tutar Int,odendi Boolean)") #kişilerin ödemesi gereken tutar ve harcamalar

        self.curs.execute("CREATE TABLE IF NOT EXISTS paketler (paketAdi Text PRIMARY KEY,tutar Int)") #kişilerin ödemesi gereken tutar ve harcamalar

        rezTakip_ui.aktiviteListele_pushButton.clicked.connect(lambda :
            self.sqlden_cagir_tabloya_dok(TABLO_AKTIVITELER,rezTakip_ui.aktivite_tableWidget))
        rezTakip_ui.rezervasyonListele_pushButton.clicked.connect(lambda :
            self.sqlden_cagir_tabloya_dok(TABLO_REZERVASYONLAR,rezTakip_ui.rezervasyon_tableWidget))
        rezTakip_ui.musterilerListele_pushButton.clicked.connect(lambda :
            self.sqlden_cagir_tabloya_dok(TABLO_MUSTERILER,rezTakip_ui.musteri_tableWidget) )

        rezTakip_ui.rezYap_pushButton.clicked.connect(lambda :
            self.satir_ekle(rezTakip_ui.rezervasyon_tableWidget))
        rezTakip_ui.aktEkle_pushButton.clicked.connect(lambda :
            self.satir_ekle(rezTakip_ui.aktivite_tableWidget))
        rezTakip_ui.musEkle_pushButton.clicked.connect(lambda :
            self.satir_ekle(rezTakip_ui.musteri_tableWidget))

        rezTakip_ui.oteldenRezGetir_pushButton.clicked.connect(lambda : self.rezervasyon_oteladi_SCTD(
                TABLO_REZERVASYONLAR,
                rezTakip_ui.rezervasyon_tableWidget,
                rezTakip_ui.musteri_tableWidget.item(rezTakip_ui.musteri_tableWidget.currentRow(),0))
                )

        rezTakip_ui.aktivitedenRezGetir_pushButton.clicked.connect(lambda : self.rezervasyon_aktivite_SCTD(
                TABLO_REZERVASYONLAR,
                rezTakip_ui.rezervasyon_tableWidget,
                rezTakip_ui.aktivite_tableWidget.item(rezTakip_ui.aktivite_tableWidget.currentRow(),0))
                )

        rezTakip_ui.UygunRezAra_pushButton.clicked.connect(lambda : self.rezervasyon_tarih_SCTD(
                TABLO_REZERVASYONLAR,
                rezTakip_ui.rezervasyon_tableWidget,
                rezTakip_ui.giris_dateEdit.date().toPyDate(),
                rezTakip_ui.cikis_dateEdit.date().toPyDate())
                )
        rezTakip_ui.Ara_pushButton.clicked.connect(self.rezervasyon_ara)


        # context menu
        rezTakip_ui.aktivite_tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        rezTakip_ui.aktivite_tableWidget.customContextMenuRequested.connect(self.show_context_menu)
        rezTakip_ui.rezervasyon_tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        rezTakip_ui.rezervasyon_tableWidget.customContextMenuRequested.connect(self.show_context_menu)
        rezTakip_ui.musteri_tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        rezTakip_ui.musteri_tableWidget.customContextMenuRequested.connect(self.show_context_menu)
    

    def rezervasyon_ara(self):
        aranan=rezTakip_ui.Arama_lineEdit.text()
        self.curs.execute(f"SELECT * FROM {TABLO_REZERVASYONLAR} WHERE (adSoyad LIKE '%{aranan}%') or (telefon LIKE '%{aranan}%')")
        data=self.curs.fetchall()
        self.tabloya_dok(rezTakip_ui.rezervasyon_tableWidget,data)
        
    def satir_ekle(self, tablo):
        global otelAdi_rezervasyon_combo_data , aktivite_rezervasyon_combo_data
        row = tablo.rowCount()
        son = tablo.columnCount()
        # Check if the row index is within the valid range
        if row >= 0:
            # Check if the widget in the specified cell is None
            widget = tablo.cellWidget(row-1, son)


            # if widget is not None and isinstance(widget, QPushButton) and widget.text() != "Kaydet":
            sender_button=rezTakip_main_window.sender()
            if isinstance(widget, CustomSpinBox):
                uyari_ver("Lütfen önce mevcut satırı kaydedin", 5000)
            elif sender_button.text() == "Rezervasyon Yap" or sender_button.text() == "Aktivite Ekle" or sender_button.text() == "Müşteri Ekle":
                tablo.setRowCount(row + 1)
                tablo.setItem(row, 0, QTableWidgetItem(""))
                tablo.item(row,0).setBackground(QColor(255,0,0))
                v_scroll_bar = tablo.verticalScrollBar()
                v_scroll_bar.setValue(v_scroll_bar.maximum())

                # Make sure to connect the correct method to the button click event
                if tablo == rezTakip_ui.aktivite_tableWidget:
                    rezTakip_ui.aktEkle_pushButton.setText("Aktivite Kaydet")
                elif tablo == rezTakip_ui.rezervasyon_tableWidget:
                    rezTakip_ui.rezYap_pushButton.setText("Rezervasyon Kaydet")

                    aktivite_combo=CustomComboBox()
                    aktivite_combo.addItem("--")
                    aktivite_combo.addItems(aktivite_rezervasyon_combo_data)
                    aktivite_combo.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")
                    tablo.setCellWidget(row, 1, aktivite_combo)

                    otelAdi_combo=CustomComboBox()
                    otelAdi_combo.addItem("--")
                    otelAdi_combo.addItems(otelAdi_rezervasyon_combo_data)
                    otelAdi_combo.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")
                    tablo.setCellWidget(row , 2, otelAdi_combo)

                    date_edit=CustomQDateEdit()
                    date_edit.setDate(datetime.today().date())
                    date_edit.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")
                    tablo.setCellWidget(row , 4,date_edit)


                    birim_combo=CustomComboBox()
                    birim_combo.addItems(["TL","DOLAR","EURO","KART"])
                    birim_combo.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")

                    tablo.setCellWidget(row , 7,birim_combo)

                elif tablo == rezTakip_ui.musteri_tableWidget:
                    rezTakip_ui.musEkle_pushButton.setText("Müşteri Kaydet")

            else:
                if tablo==rezTakip_ui.aktivite_tableWidget:
                    result=self.aktivite_kaydet()
                    if result:
                        rezTakip_ui.aktEkle_pushButton.setText("Aktivite Ekle")
                elif tablo==rezTakip_ui.rezervasyon_tableWidget:
                    result=self.rezervasyon_kaydet()
                    if result:
                        rezTakip_ui.rezYap_pushButton.setText("Rezervasyon Yap")
                elif tablo==rezTakip_ui.musteri_tableWidget:
                    result=self.musteri_kaydet()
                    if result:
                        rezTakip_ui.musEkle_pushButton.setText("Müşteri Ekle")
        else:
            uyari_ver("Geçerli bir satır indeksi bulunamadı", 5000)

    def aktivite_kaydet(self):
        try:
            row = rezTakip_ui.aktivite_tableWidget.rowCount() - 1
            try:
                aktivite = rezTakip_ui.aktivite_tableWidget.item(row, 0).text()
            except:
                raise Exception("Aktivite adı boş bırakılamaz!!")
            try:
                tl = rezTakip_ui.aktivite_tableWidget.item(row, 1).text()
            except:
                tl=0
            try:
                dolar = rezTakip_ui.aktivite_tableWidget.item(row, 2).text()
            except:
                dolar=0
            try:
                euro = rezTakip_ui.aktivite_tableWidget.item(row, 3).text()
            except:
                euro=0
            try:
                kart = rezTakip_ui.aktivite_tableWidget.item(row, 4).text()
            except:
                kart=0
            # print(aktivite,tl,dolar,euro,kart)
            self.sql_tabloya_ekle(TABLO_AKTIVITELER, (aktivite,tl,dolar,euro,kart))
            self.sqlden_cagir_tabloya_dok(TABLO_AKTIVITELER,rezTakip_ui.aktivite_tableWidget)
            return True
        except Exception as e:
            uyari_ver(str(e))
            return False

    def rezervasyon_kaydet(self):
        try:
            row = rezTakip_ui.rezervasyon_tableWidget.rowCount() - 1
            try:
                aktivite = rezTakip_ui.rezervasyon_tableWidget.cellWidget(row, 1).currentText()
            except:
                aktivite="--"
            try:
                otel_adi = rezTakip_ui.rezervasyon_tableWidget.cellWidget(row, 2).currentText()
            except:
                otel_adi="--"
            try:
                ad_soyad = rezTakip_ui.rezervasyon_tableWidget.item(row, 3).text()
            except:
                ad_soyad="--"
            try:
                rezervasyon_tarihi = rezTakip_ui.rezervasyon_tableWidget.cellWidget(row , 4).date().toPyDate()
            except:
                rezervasyon_tarihi=datetime.today().date()
            try:
                telefon = rezTakip_ui.rezervasyon_tableWidget.item(row, 5).text()
            except:
                telefon="--"
            para_birimi=rezTakip_ui.rezervasyon_tableWidget.cellWidget(row,7).currentText()
            self.curs.execute(f"SELECT {para_birimi} FROM {TABLO_AKTIVITELER} where  aktivite = ?",(aktivite,))
            fiyat= self.curs.fetchone()
            if fiyat != None:
                fiyat=fiyat[0]
                self.sql_tabloya_ekle(f"{TABLO_REZERVASYONLAR} {INSERT_REZERVASYONLAR}", (aktivite,otel_adi,ad_soyad,rezervasyon_tarihi,telefon,fiyat,para_birimi))
                self.sqlden_cagir_tabloya_dok(TABLO_REZERVASYONLAR,rezTakip_ui.rezervasyon_tableWidget)
            else:
                uyari_ver("Bu aktivite için fiyat bulunamadı !!")
                return
            # print(aktivite,otel_adi,ad_soyad,rezervasyon_tarihi,telefon,para_birimi)
            return True
        except Exception as e:
            uyari_ver(str(e))
            return False

    def musteri_kaydet(self):
        try:
            row=rezTakip_ui.musteri_tableWidget.rowCount() - 1
            otel_adi = rezTakip_ui.musteri_tableWidget.item(row, 0).text()
            try:
                telefon = rezTakip_ui.musteri_tableWidget.item(row, 1).text()
            except:
                telefon="00000000000"

            self.sql_tabloya_ekle(f"{TABLO_MUSTERILER} {INSERT_MUSTERILER}", (otel_adi,telefon))
            self.sqlden_cagir_tabloya_dok(TABLO_MUSTERILER,rezTakip_ui.musteri_tableWidget)

            return True
        except Exception as e:
            uyari_ver(str(e))
            return False


    def fiyatlari_topla(self,otel_adi,para_birimi):
        curs=self.curs
        conn=self.conn
        sql_sorgusu = """
        SELECT SUM(fiyat) AS toplamFiyat
        FROM rezervasyonlar
        WHERE otelAdi = ? AND paraBirimi = ?
        GROUP BY otelAdi, paraBirimi
        """
        # Sorguyu çalıştır
        curs.execute(sql_sorgusu, (otel_adi, para_birimi))
        toplam_fiyat=0
        # Sonuçları al
        sonuclar = curs.fetchall()
        for sonuc, in sonuclar:
            toplam_fiyat = sonuc
        return toplam_fiyat
        # Veritabanı bağlantısını kapat

    def sql_tabloya_ekle(self,tabloadi,veri):
        try:
            curs=self.curs
            conn=self.conn
            soru_isareti_sayisi=(len(veri)*'?,').strip(",")
            curs.execute(f"INSERT INTO {tabloadi} VALUES({soru_isareti_sayisi})",(veri))
            conn.commit()
            uyari_ver("Başarıyla eklendi")
        except    sqlite3.IntegrityError:
            uyari_ver("Bu otel adı daha önceden eklendi")
        except Exception as e:
            uyari_ver(str(e))


    def rezervasyon_oteladi_SCTD(self,tablo_adi,tablo,otelAdi):
            curs=self.curs
            conn=self.conn
            # print(otelAdi)
            otelAdi=otelAdi.text() if otelAdi != None else False
            if otelAdi:
                curs.execute(f"SELECT * FROM {tablo_adi} WHERE otelAdi = ?",(otelAdi,))
                data=curs.fetchall()
                self.tabloya_dok(tablo,data)
            else:
                uyari_ver("Lütfen otel adı seçiniz")

    def rezervasyon_tarih_SCTD(self,tablo_adi,tablo,baslangic_tarihi,bitis_tarihi):
        curs=self.curs
        conn=self.conn
        curs.execute(f"SELECT * FROM {tablo_adi} WHERE rezDate BETWEEN ? AND ?",(baslangic_tarihi,bitis_tarihi))
        data=curs.fetchall()
        self.tabloya_dok(tablo,data)

    def rezervasyon_aktivite_SCTD(self,tablo_adi,tablo,aktivite):
        curs=self.curs
        conn=self.conn
        aktivite=aktivite.text() if aktivite != None else Exception("Aktivite seçiniz")
        curs.execute(f"SELECT * FROM {tablo_adi} WHERE aktivite = ?",(aktivite,))
        data=curs.fetchall()
        self.tabloya_dok(tablo,data)


    def sqlden_cagir_tabloya_dok(self,tablo_adi,tablo):
        curs=self.curs
        conn=self.conn
        curs.execute(f"SELECT * FROM {tablo_adi}")
        data=curs.fetchall()
        self.tabloya_dok(tablo,data)

    def tabloya_dok(self,tablo, satirlar):
        """
        Tabloya verilen satırları ekleyen fonksiyon.

        :param tablo: QTableWidget nesnesi
        :param satirlar: Liste içinde satırlar. Her bir satır da bir liste olmalıdır.
        """
        global otelAdi_rezervasyon_combo_data , aktivite_rezervasyon_combo_data
        curs=self.curs
        conn=self.conn
        curs.execute(f"SELECT aktivite FROM {TABLO_AKTIVITELER} ORDER BY aktivite ASC")
        aktivite_rezervasyon_combo_data=[i[0] for i in curs.fetchall()]
        curs.execute(f"SELECT otelAdi FROM {TABLO_MUSTERILER} ORDER BY otelAdi ASC")
        otelAdi_rezervasyon_combo_data=[i[0] for i in curs.fetchall()]
        tablo.setRowCount(len(satirlar))
        self.aktivite_combo_data=aktivite_rezervasyon_combo_data
        self.otelAdi_combo_data=otelAdi_rezervasyon_combo_data

        # Tabloya satırları eklemek için döngü
        if tablo == rezTakip_ui.aktivite_tableWidget:
            rezTakip_ui.tabWidget.setCurrentIndex(0)
        elif tablo == rezTakip_ui.rezervasyon_tableWidget:
            rezTakip_ui.tabWidget.setCurrentIndex(1)
        elif tablo == rezTakip_ui.musteri_tableWidget:
            rezTakip_ui.tabWidget.setCurrentIndex(2)
        totalwidth = tablo.width() if tablo.width()>1550 else 1550
        for satir_index, satir in enumerate(satirlar):

            for sutun_index, hucre_verisi in enumerate(satir):

                if tablo == rezTakip_ui.aktivite_tableWidget:

                    if sutun_index==0:
                        hucre = QTableWidgetItem(str(hucre_verisi))
                        tablo.setItem(satir_index, sutun_index, hucre)
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.4))
                        tablo.item(satir_index, sutun_index).setFlags(tablo.item(satir_index, sutun_index).flags() ^ 2) ### burayı incele

                    elif 0<sutun_index<5:
                        spin_box=CustomSpinBox()
                        spin_box.setValue(float(hucre_verisi))
                        spin_box.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")
                        tablo.setCellWidget(satir_index, sutun_index,spin_box)
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.15))
                    else:                       tablo.setColumnWidth(sutun_index, 150)

                elif tablo == rezTakip_ui.rezervasyon_tableWidget:
                    # print((hucre_verisi))
                    hucre = QTableWidgetItem(str(hucre_verisi))
                    # Tabloya ekle
                    tablo.setItem(satir_index, sutun_index, hucre)
                    if sutun_index==0:
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.1))
                        tablo.item(satir_index, sutun_index).setFlags(tablo.item(satir_index, sutun_index).flags() ^ 2) ### burayı incele
                        harcama_button = QPushButton("Harcama Ekle")
                        harcama_button.setStyleSheet("background-color: rgb(125,125,250);font-size : 14px; font-family : Times New Roman")
                        harcama_button.setMaximumWidth(150)
                        harcama_button.clicked.connect(lambda : harcamaTakip(conn,curs).harcama_ekrani_ac(Table=True))
                        tablo.setCellWidget(satir_index, sutun_index, harcama_button)

                    elif sutun_index ==1:
                        aktivite_combo=CustomComboBox()
                        aktivite_combo.setEditable(True)  # burada kaldım bunu deneme için yazdım
                        aktivite_combo.addItem("--")
                        aktivite_combo.addItems(aktivite_rezervasyon_combo_data)
                        aktivite_combo.setCurrentText(hucre_verisi)
                        aktivite_combo.currentTextChanged.connect(lambda : self.tablo_update_rezervasyon())
                        aktivite_combo.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")
                        tablo.setCellWidget(satir_index, sutun_index, aktivite_combo)
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.12))

                    elif sutun_index ==2:
                        otelAdi_combo=CustomComboBox()
                        otelAdi_combo.addItem("--")
                        otelAdi_combo.addItems(otelAdi_rezervasyon_combo_data)
                        otelAdi_combo.setCurrentText(hucre_verisi)
                        otelAdi_combo.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")
                        otelAdi_combo.currentIndexChanged.connect(lambda : self.tablo_update_rezervasyon())
                        tablo.setCellWidget(satir_index, sutun_index, otelAdi_combo)
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.12))  
                    elif sutun_index == 3:
                        hucre = QTableWidgetItem(str(hucre_verisi))
                        tablo.setItem(satir_index, sutun_index, hucre)
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.11))
                    elif sutun_index == 4 :
                        date_edit=CustomQDateEdit()
                        tarih=QDate.fromString(hucre_verisi, "yyyy-MM-dd")
                        date_edit.setDate(tarih)
                        date_edit.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")
                        date_edit.setCalendarPopup(True)
                        date_edit.dateChanged.connect(lambda : self.tablo_update_rezervasyon())
                        tablo.setCellWidget(satir_index, sutun_index,date_edit)
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.09))

                    elif sutun_index == 5:
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.08))

                    elif sutun_index==6:
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.1))
                        tablo.item(satir_index, sutun_index).setFlags(tablo.item(satir_index, sutun_index).flags() ^ 2) ### burayı incele

                    elif sutun_index==7:
                        birim_combo=CustomComboBox()
                        birim_combo.addItems(["TL","DOLAR","EURO","KART"])
                        birim_combo.setCurrentText(hucre_verisi)
                        birim_combo.setStyleSheet("background-color : rgb(235,235,235); font-size : 18px; font-family : Times New Roman")
                        birim_combo.currentIndexChanged.connect(lambda : self.tablo_update_rezervasyon())
                        tablo.setCellWidget(satir_index, sutun_index,birim_combo)
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.12))
                    elif sutun_index==8:
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.1))
                        tablo.item(satir_index, sutun_index).setFlags(tablo.item(satir_index, sutun_index).flags() ^ 2)

                    else:                       tablo.setColumnWidth(sutun_index, int(totalwidth*0.1))

                elif tablo == rezTakip_ui.musteri_tableWidget:
                    hucre = QTableWidgetItem(str(hucre_verisi))
                    tablo.setItem(satir_index, sutun_index, hucre)
                    # print(sutun_index)
                    if sutun_index==0:

                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.2))
                        tablo.item(satir_index, sutun_index).setFlags(tablo.item(satir_index, sutun_index).flags() ^ 2) ### burayı incele
                        birimler=["TL","DOLAR","EURO","KART"]
                        for add,birim in enumerate(birimler): # burada otelin ödemesi gerek toplam fiyatı yazıyoruz
                            total=self.fiyatlari_topla(str(hucre_verisi),birim)
                            hucre = QTableWidgetItem(str(total))
                            tablo.setItem(satir_index, 2+add, hucre)
                            tablo.setColumnWidth(2+add, int(totalwidth*0.1))
                            tablo.item(satir_index, 2+add).setFlags(tablo.item(satir_index, 2+add).flags() ^ 2) ### burayı incele
                        sql_sorgusu = f"""
                        SELECT SUM(ekstralar) AS toplamFiyat
                        FROM {TABLO_REZERVASYONLAR}
                        WHERE otelAdi = ? 
                        GROUP BY otelAdi
                        """
                        curs.execute(sql_sorgusu, (str(hucre_verisi),))
                        sonuclar = curs.fetchall()
                        for sonuc, in sonuclar:
                            toplam_fiyat = sonuc
                            tablo.setItem(satir_index, 6, QTableWidgetItem(str(toplam_fiyat)))
                            tablo.setColumnWidth(6, int(totalwidth*0.1))
                        
                    elif sutun_index==1:
                        tablo.setColumnWidth(sutun_index, int(totalwidth*0.2))
                        tablo.item(satir_index, sutun_index).setFlags(tablo.item(satir_index, sutun_index).flags() ^ 2) ### burayı incele

                    else:
                        tablo.setColumnWidth(sutun_index, 150)


    def show_context_menu(self):
        context_menu = QMenu(rezTakip_ui.aktivite_tableWidget)

        if rezTakip_ui.tabWidget.currentIndex()==0:
            action_delete = context_menu.addAction("Aktivite Sil")
            action_update = context_menu.addAction("Aktivite Güncelle")
            action_delete.triggered.connect(lambda : self.delete_function(TABLO_AKTIVITELER,"aktivite",rezTakip_ui.aktivite_tableWidget))
            action_update.triggered.connect(lambda : self.tablo_update_aktivite())

        elif rezTakip_ui.tabWidget.currentIndex()==1:
            action_delete = context_menu.addAction("Rezervasyon Sil")
            action_update = context_menu.addAction("Rezervasyon Güncelle")
            action_delete.triggered.connect( lambda : self.delete_function(TABLO_REZERVASYONLAR,"rezID",rezTakip_ui.rezervasyon_tableWidget))
            action_update.triggered.connect(lambda : self.tablo_update_rezervasyon())

        elif rezTakip_ui.tabWidget.currentIndex()==2:
            action_delete = context_menu.addAction("Müşteri Sil")
            action_update = context_menu.addAction("Müşteri Güncelle")
            action_delete.triggered.connect( lambda : self.delete_function(TABLO_MUSTERILER,"otelAdi",rezTakip_ui.musteri_tableWidget))
            action_update.triggered.connect(lambda : self.tablo_update_musteri())

        action_exceleYaz = context_menu.addAction("Excel'e Aktar")
        action_exceleYaz.triggered.connect(lambda : self.tabloları_excel_aktar())
        context_menu.exec_(QCursor.pos())

    def delete_function(self,tabloAdi,columnName,tablo):
        selected_item=tablo.selectedItems()[0]

        if selected_item:
            if self.checkEvent("Silme İşlemi",f"Seçili veriyi {tabloAdi}'den silmek istediğinize emin misiniz ?"):
                self.curs.execute(f"DELETE FROM {tabloAdi} WHERE {columnName} = ?", (selected_item.text(),))
                self.conn.commit()
                uyari_ver(f"{selected_item.text()} başarıyla silindi")
                self.sqlden_cagir_tabloya_dok(tabloAdi,tablo)
            else:
                uyari_ver("Silme işlemi iptal edildi")
        else:
            QMessageBox.warning(tablo, "Silme İşlemi", " seçilmedi.")

    def tablo_update_aktivite(self):
        try:
            try:    row = rezTakip_ui.aktivite_tableWidget.currentRow()
            except :
                uyari_ver("Lütfen bir satır seçiniz")
                return  ""
            aktAdi =    rezTakip_ui.aktivite_tableWidget.item(row, 0).text()
            tl =        rezTakip_ui.aktivite_tableWidget.cellWidget(row,1).value()
            dolar =     rezTakip_ui.aktivite_tableWidget.cellWidget(row,2).value()
            euro =      rezTakip_ui.aktivite_tableWidget.cellWidget(row,3).value()
            kart =      rezTakip_ui.aktivite_tableWidget.cellWidget(row,4).value()
            # print(aktAdi,tl,dolar,euro,kart)

            self.curs.execute(f"UPDATE {TABLO_AKTIVITELER} SET TL= ? , DOLAR = ?,EURO = ? , KART = ? WHERE aktivite = ?",
                (tl,dolar,euro,kart,aktAdi))
            self.conn.commit()
            uyari_ver("Başarıyla güncellendi")

        except:
            uyari_ver("!! Beklenmeyen bir hata oluştu !!")

    def tablo_update_rezervasyon(self):
        try:
            try:    row = rezTakip_ui.rezervasyon_tableWidget.currentRow()
            except :
                uyari_ver("Lütfen bir satır seçiniz")
                return  ""
            id =                        rezTakip_ui.rezervasyon_tableWidget.item(row, 0).text()
            aktAdi =                    rezTakip_ui.rezervasyon_tableWidget.cellWidget(row, 1).currentText()
            otel_adi =                  rezTakip_ui.rezervasyon_tableWidget.cellWidget(row, 2).currentText()
            ad_soyad =                  rezTakip_ui.rezervasyon_tableWidget.item(row, 3).text()
            rezervasyon_tarihi =        rezTakip_ui.rezervasyon_tableWidget.cellWidget(row , 4).date().toPyDate()
            telefon =                   rezTakip_ui.rezervasyon_tableWidget.item(row, 5).text()
            para_birimi=                rezTakip_ui.rezervasyon_tableWidget.cellWidget(row,7).currentText()
            self.curs.execute(f"SELECT {para_birimi} FROM {TABLO_AKTIVITELER} where  aktivite = ?",(aktAdi,))
            fiyat= self.curs.fetchone()

            if fiyat != None:
                fiyat=fiyat[0]
                self.curs.execute(f"UPDATE {TABLO_REZERVASYONLAR} SET aktivite = ?,otelAdi = ?,adSoyad = ?,rezDate =?,telefon = ?,fiyat = ?,paraBirimi=? WHERE rezID = ?",
                    (aktAdi,otel_adi,ad_soyad,rezervasyon_tarihi,telefon,fiyat,para_birimi,        id))
                self.conn.commit()
                uyari_ver(f"{ad_soyad} Başarıyla Güncellendi")
                rezTakip_ui.rezervasyon_tableWidget.setItem(row, 6, QTableWidgetItem(str(fiyat)))
                rezTakip_ui.rezervasyon_tableWidget.item(row,6).setFlags(rezTakip_ui.rezervasyon_tableWidget.item(row,6).flags() ^ 2)
            else:
                uyari_ver("Bu aktivite için fiyat bulunamadı !!")

        except:
            uyari_ver("!! Beklenmeyen bir hata oluştu (sql_tablo_update_rezervasyon) !!")

    def tablo_update_musteri(self):
        try:
            try:    row = rezTakip_ui.musteri_tableWidget.currentRow()
            except :
                uyari_ver("Lütfen bir satır seçiniz")
                return  ""
            otel_adi =                  rezTakip_ui.rezervasyon_tableWidget.item(row, 0).text()
            telefon =                   rezTakip_ui.rezervasyon_tableWidget.item(row, 1).text()
            self.curs.execute(f"UPDATE {TABLO_MUSTERILER} SET telefon = ? WHERE otelAdi = ?",
                    (telefon,otel_adi))
            self.conn.commit()
            uyari_ver("Başarıyla güncellendi")

        except:
            uyari_ver("!! Beklenmeyen bir hata oluştu !!")


    def tabloları_excel_aktar(self):
        try:
            aktivite_liste=[[rezTakip_ui.aktivite_tableWidget.horizontalHeaderItem(i).text() for i in range(rezTakip_ui.aktivite_tableWidget.columnCount())]]
            for satir in range(rezTakip_ui.aktivite_tableWidget.rowCount()):
                satir_liste=[]
                for sutun in range(rezTakip_ui.aktivite_tableWidget.columnCount()):
                    try:
                        if sutun==0:
                            satir_liste.append(rezTakip_ui.aktivite_tableWidget.item(satir,sutun).text())
                        else:
                            satir_liste.append(rezTakip_ui.aktivite_tableWidget.cellWidget(satir,sutun).value())
                    except:
                        satir_liste.append(0)
                aktivite_liste.append(satir_liste)

            rezervasyon_liste=[[rezTakip_ui.rezervasyon_tableWidget.horizontalHeaderItem(i).text() for i in range(rezTakip_ui.rezervasyon_tableWidget.columnCount())]]
            for satir in range(rezTakip_ui.rezervasyon_tableWidget.rowCount()):
                satir_liste=[]
                for sutun in range(rezTakip_ui.rezervasyon_tableWidget.columnCount()):
                    try:
                        if sutun==4:
                            satir_liste.append(rezTakip_ui.rezervasyon_tableWidget.cellWidget(satir,sutun).date().toPyDate())
                        elif sutun ==1 or sutun==2 or sutun==7:
                            satir_liste.append(rezTakip_ui.rezervasyon_tableWidget.cellWidget(satir,sutun).currentText())
                        else:
                            satir_liste.append(rezTakip_ui.rezervasyon_tableWidget.item(satir,sutun).text())
                    except:
                        satir_liste.append("")
                rezervasyon_liste.append(satir_liste)

            musteri_liste=[[rezTakip_ui.musteri_tableWidget.horizontalHeaderItem(i).text() for i in range(rezTakip_ui.musteri_tableWidget.columnCount())]]
            for satir in range(rezTakip_ui.musteri_tableWidget.rowCount()):
                satir_liste=[]
                for sutun in range(rezTakip_ui.musteri_tableWidget.columnCount()):
                    # print(satir,sutun)
                    try:
                        satir_liste.append(rezTakip_ui.musteri_tableWidget.item(satir,sutun).text())
                    except:
                        satir_liste.append("")
                musteri_liste.append(satir_liste)
            excele_yaz([aktivite_liste,rezervasyon_liste,musteri_liste],"Output")
        except Exception as e:
            uyari_ver(str(e))


    def checkEvent(self,title,question):
            close = QMessageBox()
            close.setIcon(QMessageBox.Warning)
            close.setWindowTitle(title)
            close.setText(question)
            close.setStandardButtons(QMessageBox.Yes | QMessageBox.Cancel)
            close = close.exec()

            if close == QMessageBox.Yes:
                return True
            else:
                return False

class harcamaTakip():
    def __init__(self,conn,cursor):

        self.harcama_main_window = QMainWindow()
        self.harcama_ui = Ui_hesap_MainWindow()
        self.harcama_ui.setupUi(self.harcama_main_window)
        self.harcama_ui.statusbar.setStyleSheet("font: 75 13pt 'Times New Roman'; background-color:rgb(200,200,200);color:rgb(100,100,0)")
        self.conn=conn
        self.curs=cursor

        self.harcama_ui.harcamaEkle_pushButton.clicked.connect(lambda : self.hazir_harcama_ekle())
        self.harcama_ui.ozelharcamaEkle_pushButton.clicked.connect(lambda : self.ozel_harcama_ekle())
        self.harcama_ui.paketlerTablosu.setContextMenuPolicy(Qt.CustomContextMenu)
        self.harcama_ui.paketlerTablosu.customContextMenuRequested.connect(self.show_context_menu)
        self.harcama_ui.hesaplarTablosu.setContextMenuPolicy(Qt.CustomContextMenu)
        self.harcama_ui.hesaplarTablosu.customContextMenuRequested.connect(self.show_context_menu_hesaplar)
    def harcama_ekrani_ac(self,Table=False):

        if Table:
            try:
                sender_button = rezTakip_main_window.sender()
                if sender_button:
                    index = rezTakip_ui.rezervasyon_tableWidget.indexAt(sender_button.pos())
                    if index.isValid():
                        row = index.row()
                        self.harcama_RID=rezTakip_ui.rezervasyon_tableWidget.item(row,0).text()
                        self.harcama_adSoyad = rezTakip_ui.rezervasyon_tableWidget.item(row,3).text()
                        self.harcama_otelAdi = rezTakip_ui.rezervasyon_tableWidget.cellWidget(row, 2).currentText()

                        self.harcama_ui.baslik_label.setText("Ad Soyad: "+self.harcama_adSoyad+"\n"+"Otel: "+self.harcama_otelAdi)
                        self.harcama_rezervasyonTarihi = rezTakip_ui.rezervasyon_tableWidget.cellWidget(row,4).date().toPyDate()
                        self.harcama_ui.Tarihbaslik_label.setText(str(self.harcama_rezervasyonTarihi))
            except: rezTakip_ui.statusbar.showMessage("! hata oluştu !",5000)

        totalwidth=self.harcama_ui.hesaplarTablosu.width()+100
        self.harcama_main_window.show()
        self.harcama_ui.hesaplarTablosu.clearContents()
        self.harcama_ui.hesaplarTablosu.setColumnWidth(0,int(totalwidth/15))
        self.harcama_ui.hesaplarTablosu.setColumnWidth(1,int(totalwidth/15))
        self.harcama_ui.hesaplarTablosu.setColumnWidth(2,int(totalwidth*3.5/15))
        self.harcama_ui.hesaplarTablosu.setColumnWidth(3,int(totalwidth*4/15))
        self.harcama_ui.hesaplarTablosu.setColumnWidth(4,int(totalwidth*4/15))
        self.harcama_ui.hesaplarTablosu.setColumnWidth(5,int(totalwidth/15))


        self.harcama_tablosu_doldur()

        self.hazir_paketler_doldur()

    def harcama_tablosu_doldur(self):
        self.curs.execute(f"SELECT rezId,HarcamaId,otelAdi,adSoyad,paketAdi,tutar,odendi FROM {TABLO_HARCAMA} Where rezID = ?",(self.harcama_RID,))
        data = self.curs.fetchall()

        self.harcama_ui.hesaplarTablosu.setRowCount(len(data))
        row=0
        for rezId,HarcamaId,otelAdi,adSoyad,paketAdi,tutar,odendi in data:

            self.harcama_ui.hesaplarTablosu.setItem(row, 0, QTableWidgetItem(str(rezId)))
            self.harcama_ui.hesaplarTablosu.setItem(row, 1, QTableWidgetItem(str(HarcamaId)))
            self.harcama_ui.hesaplarTablosu.setItem(row, 2, QTableWidgetItem(str(adSoyad)))
            self.harcama_ui.hesaplarTablosu.setItem(row, 3, QTableWidgetItem(str(paketAdi)))
            self.harcama_ui.hesaplarTablosu.setItem(row, 4, QTableWidgetItem(str(tutar)))
            check_box = QCheckBox()
            check_box.setChecked(odendi)
            self.harcama_ui.hesaplarTablosu.setCellWidget(row, 5, check_box)
            check_box.stateChanged.connect(lambda state, row=row: self.check_box_changed())
            self.harcama_ui.hesaplarTablosu.setItem(row,5,QTableWidgetItem(str(odendi)))
            if odendi:
                self.setRowColor(self.harcama_ui.hesaplarTablosu,row,QColor(75,200,120))
            else:
                self.setRowColor(self.harcama_ui.hesaplarTablosu,row,QColor(200,75,120))

            row+=1

    def check_box_changed(self):
        try:
            row=-1
            sender_box=rezTakip_main_window.sender()
            if sender_box:
                    index = self.harcama_ui.hesaplarTablosu.indexAt(sender_box.pos())
                    row=index.row()
            check_box = self.harcama_ui.hesaplarTablosu.cellWidget(row, 5)
            # print("-------------------")
            if check_box.isChecked():
                self.curs.execute(f"UPDATE {TABLO_HARCAMA} SET odendi = ? WHERE HarcamaId = ?",(True,self.harcama_ui.hesaplarTablosu.item(row,1).text()))
                self.conn.commit()
                self.setRowColor(self.harcama_ui.hesaplarTablosu,row,QColor(75,200,120))
            else:
                self.curs.execute(f"UPDATE {TABLO_HARCAMA} SET odendi = ? WHERE HarcamaId = ?",(False,self.harcama_ui.hesaplarTablosu.item(row,1).text()))
                self.conn.commit()
                self.setRowColor(self.harcama_ui.hesaplarTablosu,row,QColor(200,75,120))
        except Exception as e:
            uyari_ver(str(e))
            print(str(e))



    def hazir_paketler_doldur(self):
        self.curs.execute(f"SELECT * FROM {TABLO_PAKETLER}")
        data = self.curs.fetchall()
        self.harcama_ui.paketlerTablosu.setRowCount(len(data))
        for satir_index, satir in enumerate(data):
            for sutun_index, hucre_verisi in enumerate(satir):
                hucre = QTableWidgetItem(str(hucre_verisi))
                self.harcama_ui.paketlerTablosu.setItem(satir_index, sutun_index, hucre)

    def ozel_harcama_ekle(self):
        paketAdi=self.harcama_ui.paketEkleadi_lineEdit.text()
        paketFiyat=self.harcama_ui.paketEklefiyat_lineEdit.text()
        if paketAdi and paketFiyat:
            self.harcama_ui.paketlerTablosu.setRowCount(self.harcama_ui.paketlerTablosu.rowCount() + 1)
            row=self.harcama_ui.paketlerTablosu.rowCount()-1
            self.harcama_ui.paketlerTablosu.setItem(row, 0, QTableWidgetItem(paketAdi))
            self.harcama_ui.paketlerTablosu.setItem(row, 1, QTableWidgetItem(paketFiyat))
            self.curs.execute(f"INSERT INTO {TABLO_PAKETLER} {INSERT_PAKETLER} VALUES(?,?)",(paketAdi,paketFiyat))
            self.conn.commit()
            uyari_ver("Başarıyla eklendi")

    def show_context_menu(self):
        context_menu = QMenu(self.harcama_ui.paketlerTablosu)

        action_delete = context_menu.addAction("Paket Sil")
        action_delete.triggered.connect(lambda : self.paket_sil())

        context_menu.exec_(QCursor.pos())

    def paket_sil(self):
        selected_item=self.harcama_ui.paketlerTablosu.selectedItems()[0]

        if selected_item:

                self.curs.execute(f"DELETE FROM {TABLO_PAKETLER} WHERE paketAdi = ?", (selected_item.text(),))
                self.conn.commit()
                uyari_ver(f"{selected_item.text()} başarıyla silindi")
                self.hazir_paketler_doldur()
        else:
            QMessageBox.warning(self.harcama_ui.paketlerTablosu, "Silme İşlemi", " seçilmedi.")
            
            
    def show_context_menu_hesaplar(self):
        context_menu = QMenu(self.harcama_ui.hesaplarTablosu)

        action_delete = context_menu.addAction("Harcama Sil")
        action_delete.triggered.connect(lambda : self.paket_sil_hesaplar())

        context_menu.exec_(QCursor.pos())
        
    def paket_sil_hesaplar(self):
        selected_item=self.harcama_ui.hesaplarTablosu.selectedItems()[1]

        if selected_item:
                self.curs.execute(f"DELETE FROM {TABLO_HARCAMA} WHERE HarcamaId = ?", (selected_item.text(),))
                self.conn.commit()
                uyari_ver(f"{selected_item.text()} başarıyla silindi")
                self.harcama_tablosu_doldur()
        else:
            QMessageBox.warning(self.harcama_ui.paketlerTablosu, "Silme İşlemi", " seçilmedi.")
            
            
            
    def hazir_harcama_ekle(self):
        try:
            row = self.harcama_ui.paketlerTablosu.currentRow()
            paketAdi = self.harcama_ui.paketlerTablosu.item( row, 0).text()
            fiyat = self.harcama_ui.paketlerTablosu.item( row, 1).text()
            self.hesaplar_tablosu_ekle(paketAdi,fiyat)


        except Exception as e:
            uyari_ver(str(e))

    def ekstra_toplam_fiyat(self):
        toplam=0
        odenen=0
        kalan=0
        for i in range(self.harcama_ui.hesaplarTablosu.rowCount()):
            if self.harcama_ui.hesaplarTablosu.cellWidget(i,5).isChecked():
                odenen+=float(self.harcama_ui.hesaplarTablosu.item(i,4).text())
            else:
                kalan+=float(self.harcama_ui.hesaplarTablosu.item(i,4).text())
            toplam+=float(self.harcama_ui.hesaplarTablosu.item(i,4).text())
        self.cursor.execute(f"UPDATE {TABLO_REZERVASYONLAR} SET ekstralar = ? WHERE rezID = ?",(kalan,self.harcama_RID))
        self.conn.commit()

    def hesaplar_tablosu_ekle(self,paketAdi,fiyat):
        try:

            self.harcama_ui.hesaplarTablosu.setRowCount(self.harcama_ui.hesaplarTablosu.rowCount() + 1)
            row=self.harcama_ui.hesaplarTablosu.rowCount()-1
            self.harcama_ui.hesaplarTablosu.setItem(row, 0, QTableWidgetItem(str(self.harcama_RID)))
            self.harcama_ui.hesaplarTablosu.setItem(row, 2, QTableWidgetItem(str(self.harcama_adSoyad)))
            self.harcama_ui.hesaplarTablosu.setItem(row, 3, QTableWidgetItem(paketAdi))
            self.harcama_ui.hesaplarTablosu.setItem(row, 4, QTableWidgetItem(fiyat))
            check_box = QCheckBox()
            self.harcama_ui.hesaplarTablosu.setCellWidget(row, 5, check_box)

            self.curs.execute(f"INSERT INTO {TABLO_HARCAMA} {INSERT_HARCAMA} VALUES(?,?,?,?,?,?,?)",(self.harcama_RID,self.harcama_otelAdi,self.harcama_adSoyad,self.harcama_rezervasyonTarihi,paketAdi,fiyat,False))
            self.conn.commit()
            self.ekstra_toplam_fiyat()
            return True
        except Exception as a:
            print(str(a))
            return False

    def setRowColor(self,tablo,row, color):
        for j in range(tablo.columnCount()):
            item = tablo.item(row, j)
            item.setBackground(color)


def excele_yaz(liste,dosyaAdi):
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    from openpyxl import Workbook
    fillSari    = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Burada sarı renk seçildi
    fillYesil   = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Burada yeşil renk seçildi
    fillKirmizi = PatternFill(start_color='FF0000', end_color='CC0000', fill_type='solid')  # Burada kırmızı renk seçildi
    fillMavi    = PatternFill(start_color='66CCFF', end_color='66CCFF', fill_type='solid')  # Burada mavi renk seçildi
    workbook = Workbook()
    for index,sayfa in enumerate(liste):
        sayfa_adi = "Aktiviteler" if index==0 else "Rezervasyonlar" if index==1 else "Müşteriler"
        sheet = workbook.create_sheet(sayfa_adi,index=index)
        row=1
        sayac = 0
        for satir in sayfa:
            sutunList=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
            sutunsayisi=len(satir)
            for sutunIndeks in range(sutunsayisi):
                sheet.column_dimensions[sutunList[sutunIndeks]].width = 20  # sütun için genişlik 20
            if sayac==0: # burada yazılan satır başlık satırı ise font ve renk ayarlıyoruz
                font = Font(size=13,bold=True)  # 14 punto metin boyutu
                fill=fillMavi
            else:
                font = Font(size=10)
                fill=fillSari
            for cell,col in zip(satir,sutunList):
                sheet[f'{col}{row}'] = f'{cell}'

                if sayfa_adi=="Rezervasyonlar" and (col=="G" or col=="H") and sayac!=0:
                    sheet[f'{col}{row}'].fill= fillKirmizi

                else:
                    sheet[f'{col}{row}'].fill= fill

                sheet[f'{col}{row}'].font = font
            row+=1
            sayac+=1

    try:
        workbook.save(f'{dosyaAdi}.xlsx')
        uyari_ver(f"{dosyaAdi}.xlsx adlı dosya başarıyla oluşturuldu")
        print(f"{dosyaAdi}.xlsx adlı dosya başarıyla oluşturuldu")
    except:
        print("Excel Dosyasını kapatınız")
        uyari_ver("Excel Dosyasını kapatınız")

def uyari_ver(text,wait=3000):      rezTakip_ui.statusbar.showMessage(str(text),wait)


bugun=datetime.today().date()
if bugun.month>6:
    quit()

rezTakip_ui.giris_dateEdit.setDate(bugun-timedelta(days=1))
rezTakip_ui.cikis_dateEdit.setDate(bugun+timedelta(days=6))
rezTakip_ui.giris_dateEdit.setCalendarPopup(True)
rezTakip_ui.cikis_dateEdit.setCalendarPopup(True)
try:
    sınıf=rezTakip()
except Exception as e:
    uyari_ver(str(e))

rezTakip_ui.aktiviteListele_pushButton.click()
rezTakip_ui.musterilerListele_pushButton.click()
rezTakip_ui.rezervasyonListele_pushButton.click()
rezTakip_main_window.show()
# tabloları_excel_aktar()

sys.exit(Uygulama.exec_())